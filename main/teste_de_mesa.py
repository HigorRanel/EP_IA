"""
Nomes e Nº USP:
1. Bruno Hideo Ioneda - NUSP: 15573619
2. Guilherme Samuel Lemos Segura - NUSP: 15575611
3. Higor Ranel Viani Lopes - NUSP: 15552946
4. João de Melo Fantini - NUSP: 15462550
5. Luiz Vicente Neto - NUSP: 14593054

Teste de mesa da MLP, no estilo do material de apoio da disciplina

Fixamos os pesos iniciais, escolhemos uma amostra e refazemos uma iteração
inteira na mão: forward, erro na saída, erro na oculta e atualização dos pesos.
No fim conferimos com a nossa MLP, se bater o código está certo.

O módulo confere o exemplo 2-3-2 do material, roda o teste de mesa da porta AND
com uma rede 2-2-1 e salva tudo em Excel e PDF.

Detalhe das duas notações: a nossa MLP guarda o bias separado (w0/b0); o material
junta tudo numa matriz aumentada (V e W, bias na coluna 0). As funções abaixo
traduzem de uma para a outra para exibir os resultados como v0j, vij, w0k, wjk.
"""

import os
import sys
from datetime import datetime

import numpy as np

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from mlp import MLP
import ativacoes as atv


def f4(valor):
    """Formata um número com 4 casas decimais (como no material de apoio)."""
    return f"{valor:.4f}"


def fprime(a):
    """Derivada da sigmoide: f'(a) = f(a) * (1 - f(a))."""
    return atv.derivada_sigmoid(a)


def montar_mlp_com_pesos(V, W_saida):
    """
    Cria uma MLP e injeta nela os pesos iniciais na notação do material (matrizes
    aumentadas V e W, com o bias na coluna 0). Devolve a MLP pronta.
    """
    V = np.array(V, dtype=float)
    W_saida = np.array(W_saida, dtype=float)

    p, n_mais_1 = V.shape
    n = n_mais_1 - 1
    m, p_mais_1 = W_saida.shape
    assert p_mais_1 - 1 == p, "Dimensões de V e W_saida incompatíveis."

    mlp = MLP(
        comprimento_entrada=n,
        comprimento_oculta=p,
        comprimento_saida=m,
        epocas=1,
        taxa_de_aprendizado=0.5,
        ini_pesos='xavier',
        ini_bias='zero',
    )

    # Coluna 0 é o bias; o resto são os pesos (transpostos p/ a forma da MLP).
    mlp.w0 = V[:, 0].copy()
    mlp.W = V[:, 1:].T.copy()
    mlp.b0 = W_saida[:, 0].copy()
    mlp.B = W_saida[:, 1:].T.copy()

    return mlp


def gerar_trace(V, W_saida, x, t, alpha=0.5):
    """
    Faz uma iteração de backpropagation na mão (e em paralelo pela MLP, p/
    conferir) e devolve um dicionário com as tabelas de cada passo do teste de
    mesa: forward, backward da saída, backward da oculta e pesos novos.
    """
    V = np.array(V, dtype=float)
    W_saida = np.array(W_saida, dtype=float)
    x = np.array(x, dtype=float)
    t = np.array(t, dtype=float)

    p = V.shape[0]
    n = V.shape[1] - 1
    m = W_saida.shape[0]

    # Forward pela MLP, de onde tiramos z_in, z, y_in e y.
    mlp = montar_mlp_com_pesos(V, W_saida)
    mlp.forward(x)
    z_in = np.array(mlp.z_in, dtype=float)
    z = np.array(mlp.z, dtype=float)
    y_in = np.array(mlp.y_in, dtype=float)
    y = np.array(mlp.y, dtype=float)

    # Vetores "aumentados" com o bias na posição 0 (x0 = 1, z0 = 1).
    x_aug = np.concatenate(([1.0], x))
    z_aug = np.concatenate(([1.0], z))

    forward_rows = []
    for j in range(p):
        termos = " + ".join(f"{f4(x_aug[i])}*{f4(V[j, i])}" for i in range(n + 1))
        forward_rows.append({
            'Camada': 'Oculta',
            'Símbolo': f'z_in_{j + 1}',
            'Fórmula': f'Σ x_i·v_i{j + 1}',
            'Substituindo': termos,
            'Resultado': round(float(z_in[j]), 4),
        })
        forward_rows.append({
            'Camada': 'Oculta',
            'Símbolo': f'z_{j + 1}',
            'Fórmula': f'f(z_in_{j + 1})',
            'Substituindo': f'sigmoide({f4(z_in[j])})',
            'Resultado': round(float(z[j]), 4),
        })
    for k in range(m):
        termos = " + ".join(f"{f4(z_aug[j])}*{f4(W_saida[k, j])}" for j in range(p + 1))
        forward_rows.append({
            'Camada': 'Saída',
            'Símbolo': f'y_in_{k + 1}',
            'Fórmula': f'Σ z_j·w_j{k + 1}',
            'Substituindo': termos,
            'Resultado': round(float(y_in[k]), 4),
        })
        forward_rows.append({
            'Camada': 'Saída',
            'Símbolo': f'y_{k + 1}',
            'Fórmula': f'f(y_in_{k + 1})',
            'Substituindo': f'sigmoide({f4(y_in[k])})',
            'Resultado': round(float(y[k]), 4),
        })

    # Backward na saída: δ_s_k = (t_k - y_k) · f'(y_in_k)
    delta_s = (t - y) * np.array([fprime(v) for v in y_in])

    backward_saida_rows = []
    for k in range(m):
        backward_saida_rows.append({
            'Símbolo': f'δ_s_{k + 1}',
            'Fórmula': f'(t_{k + 1} - y_{k + 1})·f\'(y_in_{k + 1})',
            'Substituindo': f'({f4(t[k])} - {f4(y[k])})·{f4(fprime(y_in[k]))}',
            'Resultado': round(float(delta_s[k]), 4),
        })
    # Δw_jk = α · δ_s_k · z_j  (j = 0 é o bias, z0 = 1)
    delta_w = np.zeros((m, p + 1))
    for k in range(m):
        for j in range(p + 1):
            delta_w[k, j] = alpha * delta_s[k] * z_aug[j]
            backward_saida_rows.append({
                'Símbolo': f'Δw_{j}{k + 1}',
                'Fórmula': f'α·δ_s_{k + 1}·z_{j}',
                'Substituindo': f'{f4(alpha)}·{f4(delta_s[k])}·{f4(z_aug[j])}',
                'Resultado': round(float(delta_w[k, j]), 4),
            })

    # Backward na oculta: δ_in_j = Σ_k δ_s_k · w_jk ; δ_h_j = δ_in_j · f'(z_in_j)
    B = W_saida[:, 1:]  # B[k, j]: peso do oculto (j+1) para a saída (k+1)
    delta_in = np.array([sum(delta_s[k] * B[k, j] for k in range(m)) for j in range(p)])
    delta_h = delta_in * np.array([fprime(v) for v in z_in])

    backward_oculta_rows = []
    for j in range(p):
        soma = " + ".join(f"{f4(delta_s[k])}*{f4(B[k, j])}" for k in range(m))
        backward_oculta_rows.append({
            'Símbolo': f'δ_in_{j + 1}',
            'Fórmula': f'Σ δ_s_k·w_{j + 1}k',
            'Substituindo': soma,
            'Resultado': round(float(delta_in[j]), 4),
        })
        backward_oculta_rows.append({
            'Símbolo': f'δ_h_{j + 1}',
            'Fórmula': f'δ_in_{j + 1}·f\'(z_in_{j + 1})',
            'Substituindo': f'{f4(delta_in[j])}·{f4(fprime(z_in[j]))}',
            'Resultado': round(float(delta_h[j]), 4),
        })
    # Δv_ij = α · δ_h_j · x_i  (i = 0 é o bias, x0 = 1)
    delta_v = np.zeros((p, n + 1))
    for j in range(p):
        for i in range(n + 1):
            delta_v[j, i] = alpha * delta_h[j] * x_aug[i]
            backward_oculta_rows.append({
                'Símbolo': f'Δv_{i}{j + 1}',
                'Fórmula': f'α·δ_h_{j + 1}·x_{i}',
                'Substituindo': f'{f4(alpha)}·{f4(delta_h[j])}·{f4(x_aug[i])}',
                'Resultado': round(float(delta_v[j, i]), 4),
            })

    # Atualização: peso novo = peso antigo + correção.
    V_novo = V + delta_v
    W_novo = W_saida + delta_w

    pesos_rows = []
    for j in range(p):
        for i in range(n + 1):
            pesos_rows.append({
                'Peso': f'v_{i}{j + 1}',
                'Antigo': round(float(V[j, i]), 4),
                'Δ': round(float(delta_v[j, i]), 4),
                'Novo': round(float(V_novo[j, i]), 4),
            })
    for k in range(m):
        for j in range(p + 1):
            pesos_rows.append({
                'Peso': f'w_{j}{k + 1}',
                'Antigo': round(float(W_saida[k, j]), 4),
                'Δ': round(float(delta_w[k, j]), 4),
                'Novo': round(float(W_novo[k, j]), 4),
            })

    # A MLP chega nos mesmos pesos finais? Remontamos as matrizes p/ comparar.
    mlp.backpropagate(x, t, mlp.z_in, mlp.z, mlp.y_in, mlp.y)
    V_mlp = np.column_stack([mlp.w0, mlp.W.T])
    W_mlp = np.column_stack([mlp.b0, mlp.B.T])
    bate = (np.allclose(V_mlp, V_novo, atol=1e-9)
            and np.allclose(W_mlp, W_novo, atol=1e-9))

    parametros_rows = [
        {'Item': 'Nº de entradas (n)', 'Valor': n},
        {'Item': 'Nº de neurônios ocultos (p)', 'Valor': p},
        {'Item': 'Nº de neurônios de saída (m)', 'Valor': m},
        {'Item': 'Taxa de aprendizado (α)', 'Valor': alpha},
        {'Item': 'Função de ativação', 'Valor': 'sigmoide'},
        {'Item': 'Entrada x', 'Valor': str(list(x))},
        {'Item': 'Alvo t', 'Valor': str(list(t))},
    ]

    return {
        'parametros': parametros_rows,
        'forward': forward_rows,
        'backward_saida': backward_saida_rows,
        'backward_oculta': backward_oculta_rows,
        'pesos': pesos_rows,
        'confere_com_mlp': bool(bate),
        # valores brutos, úteis para o texto humanizado
        'bruto': {
            'x': x, 't': t, 'alpha': alpha, 'n': n, 'p': p, 'm': m,
            'z_in': z_in, 'z': z, 'y_in': y_in, 'y': y,
            'delta_s': delta_s, 'delta_in': delta_in, 'delta_h': delta_h,
            'V': V, 'W_saida': W_saida, 'V_novo': V_novo, 'W_novo': W_novo,
            'erro': 0.5 * float(np.sum((t - y) ** 2)),
        },
    }


def salvar_excel(trace, caminho):
    """
    Grava o teste de mesa numa planilha em que a matematica é feita pelo próprio
    Excel. Cada valor é uma fórmula viva que referencia as células dos pesos e
    entradas (mude um peso e tudo recalcula). A sigmoide vira =1/(1+EXP(-a)) e a
    derivada é reaproveitada como f'(a) = saída·(1 - saída).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    b = trace['bruto']
    V, W, x, t = b['V'], b['W_saida'], b['x'], b['t']
    alpha, n, p, m = b['alpha'], b['n'], b['p'], b['m']

    wb = Workbook()
    ws = wb.active
    ws.title = 'Teste de Mesa (fórmulas)'
    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 50

    azul = PatternFill('solid', fgColor='4472C4')
    cinza = PatternFill('solid', fgColor='D9E1F2')

    estado = {'row': 1}
    addr = {}

    def titulo(texto):
        r = estado['row']
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        c = ws.cell(row=r, column=1, value=texto)
        c.font = Font(bold=True, size=13)
        estado['row'] += 2

    def secao(texto):
        r = estado['row']
        for col in (1, 2, 3):
            ws.cell(row=r, column=col).fill = azul
        c = ws.cell(row=r, column=1, value=texto)
        c.font = Font(bold=True, color='FFFFFF')
        estado['row'] += 1

    def put(label, valor, descricao='', key=None, guardar=True, destaque=False):
        r = estado['row']
        ca = ws.cell(row=r, column=1, value=label)
        cb = ws.cell(row=r, column=2, value=valor)
        cb.number_format = '0.0000'
        if descricao:
            ws.cell(row=r, column=3, value=descricao)
        if destaque:
            ca.font = Font(bold=True)
            cb.font = Font(bold=True)
            cb.fill = cinza
        if guardar:
            addr[key or label] = f'B{r}'
        estado['row'] += 1

    def soma(parcelas):
        """Monta uma fórmula de soma de produtos: '=a*b+c*d+...'."""
        return '=' + '+'.join(parcelas)

    titulo(f'TESTE DE MESA — PORTA AND (rede {n}-{p}-{m}, ativação sigmoide)')

    secao('PARÂMETROS')
    put('alpha', float(alpha), 'taxa de aprendizado (α)')

    secao('ENTRADAS (x0 = 1 é o bias)')
    put('x0', 1.0, 'bias de entrada (sempre 1)')
    for i in range(1, n + 1):
        put(f'x{i}', float(x[i - 1]), f'atributo x{i} da amostra')

    secao('ALVO')
    for k in range(1, m + 1):
        put(f't{k}', float(t[k - 1]), f'saída desejada do neurônio {k}')

    secao('PESOS INICIAIS — CAMADA OCULTA (V)')
    for j in range(1, p + 1):
        for i in range(0, n + 1):
            put(f'v{i}{j}', float(V[j - 1, i]),
                f'peso x{i} → oculto {j}' + (' (bias)' if i == 0 else ''))

    secao('PESOS INICIAIS — CAMADA DE SAÍDA (W)')
    for k in range(1, m + 1):
        for j in range(0, p + 1):
            put(f'w{j}{k}', float(W[k - 1, j]),
                f'peso z{j} → saída {k}' + (' (bias)' if j == 0 else ''))

    secao('PASSO 1 — FORWARD (a rede calcula a saída)')
    for j in range(1, p + 1):
        put(f'z_in{j}',
            soma(f'{addr[f"x{i}"]}*{addr[f"v{i}{j}"]}' for i in range(n + 1)),
            f'z_in{j} = Σ x_i · v_i{j}')
        put(f'z{j}', f'=1/(1+EXP(-{addr[f"z_in{j}"]}))', f'z{j} = sigmoide(z_in{j})')
    put('z0', 1.0, 'bias para a camada de saída (sempre 1)')
    for k in range(1, m + 1):
        put(f'y_in{k}',
            soma(f'{addr[f"z{j}"]}*{addr[f"w{j}{k}"]}' for j in range(p + 1)),
            f'y_in{k} = Σ z_j · w_j{k}')
        put(f'y{k}', f'=1/(1+EXP(-{addr[f"y_in{k}"]}))',
            f'y{k} = sigmoide(y_in{k})', destaque=True)
    put('E', '=0.5*(' + '+'.join(
        f'({addr[f"t{k}"]}-{addr[f"y{k}"]})^2' for k in range(1, m + 1)) + ')',
        'erro quadrático da amostra: ½·Σ(t−y)²')

    secao('PASSO 2 — BACKWARD (camada de saída)')
    for k in range(1, m + 1):
        put(f"f'(y_in{k})", f'={addr[f"y{k}"]}*(1-{addr[f"y{k}"]})',
            "f'(y_in) = y·(1−y)", key=f'fly{k}')
        put(f'δ_s{k}', f'=({addr[f"t{k}"]}-{addr[f"y{k}"]})*{addr[f"fly{k}"]}',
            f'δ_s{k} = (t{k} − y{k}) · f\'(y_in{k})', key=f'ds{k}', destaque=True)
    for k in range(1, m + 1):
        for j in range(0, p + 1):
            put(f'Δw{j}{k}',
                f'={addr["alpha"]}*{addr[f"ds{k}"]}*{addr[f"z{j}"]}',
                f'Δw{j}{k} = α · δ_s{k} · z{j}', key=f'dw{j}{k}')

    secao('PASSO 3 — BACKWARD (camada oculta)')
    for j in range(1, p + 1):
        put(f'δ_in{j}',
            soma(f'{addr[f"ds{k}"]}*{addr[f"w{j}{k}"]}' for k in range(1, m + 1)),
            f'δ_in{j} = Σ δ_s_k · w{j}k', key=f'din{j}')
        put(f"f'(z_in{j})", f'={addr[f"z{j}"]}*(1-{addr[f"z{j}"]})',
            "f'(z_in) = z·(1−z)", key=f'flz{j}')
        put(f'δ_h{j}', f'={addr[f"din{j}"]}*{addr[f"flz{j}"]}',
            f'δ_h{j} = δ_in{j} · f\'(z_in{j})', key=f'dh{j}', destaque=True)
    for j in range(1, p + 1):
        for i in range(0, n + 1):
            put(f'Δv{i}{j}',
                f'={addr["alpha"]}*{addr[f"dh{j}"]}*{addr[f"x{i}"]}',
                f'Δv{i}{j} = α · δ_h{j} · x{i}', key=f'dv{i}{j}')

    secao('PASSO 4 — ATUALIZAÇÃO DOS PESOS (regra delta)')
    for j in range(1, p + 1):
        for i in range(0, n + 1):
            put(f'v{i}{j}_novo', f'={addr[f"v{i}{j}"]}+{addr[f"dv{i}{j}"]}',
                f'v{i}{j} + Δv{i}{j}', guardar=False)
    for k in range(1, m + 1):
        for j in range(0, p + 1):
            put(f'w{j}{k}_novo', f'={addr[f"w{j}{k}"]}+{addr[f"dw{j}{k}"]}',
                f'w{j}{k} + Δw{j}{k}', guardar=False)

    estado['row'] += 1
    nota = ws.cell(row=estado['row'], column=1,
                   value='Abra no Excel/LibreOffice: a coluna B é recalculada a '
                         'partir dos pesos e entradas (mude um peso e veja tudo mudar).')
    nota.font = Font(italic=True, color='808080')
    ws.merge_cells(start_row=estado['row'], start_column=1,
                   end_row=estado['row'], end_column=3)

    wb.save(caminho)
    print(f"Excel (com fórmulas vivas) do teste de mesa salvo em: {caminho}")


def _equacao(linha):
    """Formata uma linha de trace como 'símbolo = substituição = resultado'."""
    subs = linha['Substituindo'].replace('*', '×')
    return f"{linha['Símbolo']} = {subs} = {f4(linha['Resultado'])}"


def _render_pdf(linhas, caminho):
    """Desenha uma lista de (estilo, texto) num PDF paginado, via matplotlib.
    Estilos: 'titulo', 'secao', 'normal' e 'espaco'."""
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages

    dy = 0.0165
    pdf = PdfPages(caminho)
    fig = plt.figure(figsize=(8.27, 11.69))  # A4 retrato
    estado = {'y': 0.95}

    def nova_pagina():
        pdf.savefig(fig)
        plt.close(fig)
        nova = plt.figure(figsize=(8.27, 11.69))
        return nova

    for estilo, texto in linhas:
        if estado['y'] < 0.06:
            fig = nova_pagina()
            estado['y'] = 0.95
        y = estado['y']
        if estilo == 'titulo':
            fig.text(0.5, y, texto, ha='center', family='monospace',
                     fontsize=14, fontweight='bold')
            estado['y'] -= dy * 2.4
        elif estilo == 'secao':
            estado['y'] -= dy * 0.6
            fig.text(0.07, estado['y'], texto, family='monospace',
                     fontsize=11, fontweight='bold', color='#1F3B6F')
            estado['y'] -= dy * 1.5
        elif estilo == 'espaco':
            estado['y'] -= dy
        else:
            fig.text(0.085, y, texto, family='monospace', fontsize=8.5)
            estado['y'] -= dy

    pdf.savefig(fig)
    plt.close(fig)
    pdf.close()
    print(f"PDF do teste de mesa salvo em: {caminho}")


def salvar_pdf(trace, caminho, titulo='Teste de mesa da MLP — Porta AND'):
    """Monta o PDF do teste de mesa no estilo do material: configuração, os
    quatro passos com as contas substituídas e a conferência final."""
    b = trace['bruto']
    n, p, m = b['n'], b['p'], b['m']
    V, W = b['V'], b['W_saida']

    linhas = [('titulo', titulo)]

    linhas.append(('secao', 'Configuração da rede'))
    linhas.append(('normal',
                   f'Arquitetura: {n} entradas — {p} ocultos — {m} saída(s)   |   '
                   f'ativação: sigmoide   |   α = {f4(b["alpha"])}'))
    linhas.append(('normal',
                   f'Amostra:  x = {[f4(v) for v in b["x"]]}  (x0 = 1)      '
                   f'alvo  t = {[f4(v) for v in b["t"]]}'))
    linhas.append(('normal', 'Pesos iniciais da camada oculta (V):'))
    for j in range(p):
        linhas.append(('normal', '   ' + '    '.join(
            f'v{i}{j + 1}={f4(V[j, i])}' for i in range(n + 1))))
    linhas.append(('normal', 'Pesos iniciais da camada de saída (W):'))
    for k in range(m):
        linhas.append(('normal', '   ' + '    '.join(
            f'w{jj}{k + 1}={f4(W[k, jj])}' for jj in range(p + 1))))

    linhas.append(('espaco', ''))
    linhas.append(('secao', 'Passo 1 — Forward (a rede calcula sua saída)'))
    for r in trace['forward']:
        linhas.append(('normal', _equacao(r)))
    linhas.append(('normal',
                   f'Erro quadrático da amostra:  E = ½·Σ(t − y)² = {f4(b["erro"])}'))

    linhas.append(('espaco', ''))
    linhas.append(('secao', 'Passo 2 — Retropropagação na camada de saída'))
    for r in trace['backward_saida']:
        linhas.append(('normal', _equacao(r)))

    linhas.append(('espaco', ''))
    linhas.append(('secao', 'Passo 3 — Retropropagação na camada oculta'))
    for r in trace['backward_oculta']:
        linhas.append(('normal', _equacao(r)))

    linhas.append(('espaco', ''))
    linhas.append(('secao', 'Passo 4 — Atualização dos pesos (regra delta)'))
    for r in trace['pesos']:
        linhas.append(('normal',
                       f"{r['Peso']:<7}: {f4(r['Antigo'])}  +  ({f4(r['Δ'])})  "
                       f"=  {f4(r['Novo'])}"))

    linhas.append(('espaco', ''))
    linhas.append(('secao', 'Conferência'))
    conf = 'SIM' if trace['confere_com_mlp'] else 'NÃO'
    linhas.append(('normal',
                   f'Os pesos finais calculados batem com a implementação MLP:  {conf}'))

    _render_pdf(linhas, caminho)


def verificar_exemplo_do_material():
    """Reproduz o exemplo 2-3-2 do material e confere se a MLP devolve os mesmos
    números do PDF — uma prova de corretude independente da porta AND."""
    # Pesos exatamente como no material (Seção 3 do PDF).
    V = [
        [-0.1, 0.1, -0.1],
        [-0.1, 0.1, 0.1],
        [0.1, -0.1, -0.1],
    ]
    W_saida = [
        [-0.1, 0.1, 0.0, 0.1],
        [0.1, -0.1, 0.1, -0.1],
    ]
    x = [1.0, 1.0]
    t = [1.0, 0.0]

    trace = gerar_trace(V, W_saida, x, t, alpha=0.5)
    b = trace['bruto']

    print("\n" + "=" * 60)
    print("CONFERÊNCIA COM O MATERIAL DE APOIO (rede 2-3-2)".center(60))
    print("=" * 60)
    print(f"z = {[f4(v) for v in b['z']]}   (PDF: 0.4750, 0.5250, 0.4750)")
    print(f"y = {[f4(v) for v in b['y']]}   (PDF: 0.4988, 0.5144)")
    print(f"δ_saída = {[f4(v) for v in b['delta_s']]}   (PDF: 0.1253, -0.1285)")
    print(f"δ_oculta = {[f4(v) for v in b['delta_h']]}   (PDF: 0.0063, -0.0032, 0.0063)")
    print(f"w01 novo = {f4(b['W_novo'][0, 0])}   (PDF: -0.0373)")
    print(f"v11 novo = {f4(b['V_novo'][0, 1])}   (PDF:  0.1032)")

    # Conferência automática contra os valores do PDF.
    esperado_z = np.array([0.4750, 0.5250, 0.4750])
    esperado_y = np.array([0.4988, 0.5144])
    ok = (np.allclose(b['z'], esperado_z, atol=1e-3)
          and np.allclose(b['y'], esperado_y, atol=1e-3)
          and trace['confere_com_mlp'])
    print(f"\nResultado: {'BATE com o material de apoio.' if ok else 'NÃO bate — revisar.'}")
    return trace


def main():
    """Confere o exemplo do material e gera o teste de mesa da AND, excel, PDF e
    texto, salvando tudo em ./Saidas/execucao_teste_<anomesdia_horaminseg>."""
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    saidas = os.path.join(BASE_DIR, 'Saidas', f'execucao_teste_{timestamp}')
    os.makedirs(saidas, exist_ok=True)

    import loggers.writer as writer_mod

    def _writer_pasta_fixa(self, diretorio_saida="./Saidas"):
        self.diretorio_saida = diretorio_saida
        self.pasta_atual = saidas
        os.makedirs(self.pasta_atual, exist_ok=True)

    writer_mod.Writer.__init__ = _writer_pasta_fixa

    verificar_exemplo_do_material()

    # Rede 2-2-1, pesos iniciais simples
    V = [
        [0.1, 0.2, -0.2],
        [-0.1, 0.3, 0.1], 
    ]
    W_saida = [
        [0.2, -0.3, 0.1],
    ]
    x = [1.0, 1.0]
    t = [1.0]

    trace = gerar_trace(V, W_saida, x, t, alpha=0.5)

    print("\n" + "=" * 60)
    print("TESTE DE MESA DA PORTA AND (rede 2-2-1)".center(60))
    print("=" * 60)
    b = trace['bruto']
    print(f"z = {[f4(v) for v in b['z']]}")
    print(f"y = {[f4(v) for v in b['y']]}  (alvo t = {f4(t[0])})")
    print(f"δ_saída = {[f4(v) for v in b['delta_s']]}")
    print(f"δ_oculta = {[f4(v) for v in b['delta_h']]}")
    print(f"Cálculo no papel confere com a MLP? "
          f"{'SIM' if trace['confere_com_mlp'] else 'NÃO'}")

    salvar_excel(trace, os.path.join(saidas, 'teste_de_mesa_AND.xlsx'))
    salvar_narrativa(trace, os.path.join(saidas, 'teste_de_mesa_AND.md'))
    salvar_pdf(trace, os.path.join(saidas, 'teste_de_mesa_AND.pdf'))


if __name__ == '__main__':
    main()
